from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Catchment,Person,Project
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
import json
from openpyxl import load_workbook
from io import BytesIO, StringIO
import requests
import pyrebase
import firebase_admin
from firebase_admin import credentials
import csv
from openpyxl import Workbook

config = {
     "apiKey": "AIzaSyDGrA3lZhidUwBRar9zWiS4vXzgja0XTXQ",
  "authDomain": "kwathu-b7b68.firebaseapp.com",
    "databaseURL" : "https://kwathu-b7b68.appspot.com",
  "projectId": "kwathu-b7b68",
  "storageBucket": "kwathu-b7b68.appspot.com",
  "messagingSenderId": "284956368045",
  "appId": "1:284956368045:web:da970effd3cb6db08a3a24",
  "measurementId": "G-9QEWDV7B17",
  "type": "service_account",
  "private_key_id": "6a2bda240a06dd09d91e6e4bd4f0378bf697d602",
  "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQDof/AbY3UW9aIR\nVWen6JqW9J4A9fbBRGX1+B4ZBq6YTfCZEIAoK3hCpqn9gK4A/J5TkO7z+b/sPYlT\ndrDQzEviWja/hxOfTc/s3XBUqs1IuVyLsDwvf4gvpk/R4OlFgaHtZSImzQL5vvaw\nLh+V079lg6xlDNtuNSnhsVSDBQIj2+HrIYd/699j2dqWTEFgmWtkYcsf/TiE2oM+\nxZ3uRAMKjRSNPEg7zr+8UAZ8AGJn9WCBXEdf5mO/Xs0fXM8owQrdBpCuFpVk3kgn\nKq8lA+kkA5KHc49unJ7gX4Rb666/JCcPSq0zyXpRf0GSeozkKGUPYMlB15KrComP\ndsQt2QNPAgMBAAECggEAbcETfkJloFiFUMxx2dRWwcMpIYf7G8InpcUipCrA2foV\nBZTCAvR+WWxm29pWifW24fkNrFOuU527CZECqBaBDhILPT/NNgscCBOCtx1IOjfy\npaikyMFZ7zCCBrA3Po4S/EedJxbpfC83MfUcCPUEKa5OgAdhQLXNFxNIs272UylP\nmJCoCJc6TOP+xmwuru2uDQayXMYsECoxDMGM7vQbXg+GgYuhaStwOyVWZ6mh4ght\nIx3HPQ6Ccbrr3gwvTlogwDAKy7CaBRv6P/8w2DaNJmY0Fsz/PtO2KpbnyeeRWLZi\nktuUt3JpDWMfI4mQOuzRULNdHkz3D5mtJMRS58PZrQKBgQD75FJ3PsUx2cCQxMvv\noR46xqqzKLPRNTUdgDumYv9CHzfyW2tnrSrbUxKLE9MTtIo40cUtVVb8QsRGw+ot\nRvKh6CK6Q5CgxVipYKP7Z3u4UbBpbUAFE6afmy/6Y1th66g6hVuwucsC6V7cSIUN\nUlNIs7VShIX7h22bq89YGWbjowKBgQDsSqbCsXhokYXeQ41s94K6dqCcr1RJRFDb\nUQSG3eMwAaE3FnZuL26lIia3wkFtVjSC5ekX3+frUWooev+8C+eUtwm4MM/X2q6q\nBFbK6iI9GgoGOFmDTGTCQx7cuWD+1IVI1EXImQgSy7dAmwqjWYn5XTDbv316BhyP\nouAaT+08ZQKBgQCL33WLkeeejvRHb4BvSGCo8rjHU8rGWW1tPE/jTQXBieRfSDSU\nFay+rZqWcCO9xXsboiCD0+fBnLGxOglHqyYAkg72r54YKH3bUDa+2+/voiz73ha2\nuvrDlkFGIPYVoe6A0JzhvcnljAGJARwhEZ39uQzvDXf4/HL5BqbUnQZtUQKBgCI8\nztz3ChNcwqDp+NwJIRATkEIjxXq4Q18D5jwEep7CHbH1t0NP7CBRSnrl0sUc8I7m\n8VdcQiZ+rXsgF14P+4Y431eo+vSz4wsPhePw+PURiNJN5+p2b0MWpqok0kEaWhWr\nbFEcm2bdzbe2v1vb5XoUXrf4jx3XNNw7JW2qlx1hAoGBAJNKvuU0YEBgrqTadkfq\nm3YL8CuYBS4HNzocbm95GR4wH67YjdJgRqwM9WfxkiSn9nYKXkEvJJG4TtnoDJ+n\nr3bCKli7QuVAjvc0MlCh+KrxpGbIdnz/3SC8eDpJQdVQGPDlzLJpiHjOlxkEejGB\nOhOcRtrCtD7HURGGx7Ps31qY\n-----END PRIVATE KEY-----\n",
  "client_email": "firebase-adminsdk-mtu1r@kwathu-b7b68.iam.gserviceaccount.com",
  "client_id": "108729172439378885152",
  "auth_uri": "https://accounts.google.com/o/oauth2/auth",
  "token_uri": "https://oauth2.googleapis.com/token",
  "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
  "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-mtu1r%40kwathu-b7b68.iam.gserviceaccount.com",
  "universe_domain": "googleapis.com"
}

path = r'app/kwathu-b7b68-firebase-adminsdk-mtu1r-6a2bda240a.json'
firebase = pyrebase.initialize_app(config)
if not firebase_admin._apps:
    cred = credentials.Certificate(path)
    firebase_admin.initialize_app(cred)

# def register(name : str,xpath : str):
#     try:   
#         db = firestore.client()
#         dbDoc = db.collection("miso").document('data').collection('processed').document(name)
#         dbDoc.set({"name" : name, "path" : xpath})
#         print(True)
        
#     except Exception as e :
#         print(e)
#         return False   
def download(path : str, name : str):
    file = requests.get(path)
    if file.status_code == 200:
        file_content = BytesIO(file.content)
        csv_check = checkIfCSV(path=name)
        if csv_check:
            return csv_to_xlsx(file_content)
        else:
            file_workbook = load_workbook(file_content)
            return file_workbook
    else:
        return False

def upload(my_path : str,name : str) :
    try:        
        storage = firebase.storage()
        target_path = f"miso/processed/" + name
        storage.child(target_path).put(my_path)
        download_url = storage.child(target_path).get_url(None)
        return download_url
    except Exception as e:
        print(e)
        return False
def create_path(my_workbook):
    output = BytesIO()
    my_workbook.save(output)
    output.seek(0)
    return output

def load_reference()->list:
    reference = Person.objects.all()
    refs = list(map(lambda person : [person.phoneNumber,person.form_number],reference))
    return refs

def load_external_reference(path : str)->list:
    myworkbook = load_workbook(filename = path)
    return [rows for rows in myworkbook.active.iter_rows()]
  
@csrf_exempt
def clear_users(request):
    if request.method == "OPTIONS":
        response = HttpResponse()
        response['Access-Control-Allow-Origin'] = '' 
        response['Access-Control-Allow-Methods'] = 'POST'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response
    elif request.method == "POST":
        try:
           users = Person.objects.all()
           for user in users:
               user.delete()
        except Exception as e:
            print(e)
            return toJsonResponse({"status" : False,"message" : f"something went wrong \n {e}"})




def add_phone_numbers(target_path : str,reference : list,output_path : str = "nkhoma",filename : str = ""):
    #my_workbook = load_workbook(filename = target_path)
    my_workbook = download(target_path,name=filename)
    my_rows = [rows for rows in my_workbook.active.iter_rows()]
    header_target = 0
    header_formnumber = 1
    header_target_final = [ref.value for ref in my_rows[0]].index("Phone Number") 
    if not header_target_final:
        header_target_final = 15
    formnumber_target_final = [ref.value for ref in my_rows[0]].index("form_number")

    
    for phone in enumerate(my_rows):
            if phone[0] != 0:
                target_row = phone[1]
                target_cell = target_row[header_target_final]
                for value in reference:
                    if str(value[header_formnumber]) == str(target_row[formnumber_target_final].value):
                        target_cell.value = value[header_target]  
                    
            
    xpath =   create_path(my_workbook)
    url = upload(xpath,output_path)
    # register(output_path,url)
    
    print("done")
    # my_workbook.save(output_path)
    return url
    
    
def loadJsonData(data):
    return json.loads(data.body.decode())

def toJsonResponse(data)->HttpResponse:
    json_data = json.dumps(data)   
    response = HttpResponse(json_data, content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename="data.json"'
    return response

@csrf_exempt
def assignFile(request):
    if request.method == "OPTIONS":
        response = HttpResponse()
        response['Access-Control-Allow-Origin'] = '' 
        response['Access-Control-Allow-Methods'] = 'POST'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response
    elif request.method == "POST":
        try:
            print(request.body.decode())
            data = loadJsonData(request)
            name = data.get("name")
            path = data.get("path")
            print(0)
            reference = load_reference()
            if name and path and reference:
                uri = add_phone_numbers(target_path = path,reference=reference,output_path=name, filename= name)
                return toJsonResponse({"status" : True,"message" : f"{uri}"})
            else:
                return toJsonResponse({"status" : False,"message" : f"info missing"})
        except Exception as e:
            print(e)
            return toJsonResponse({"status" : False,"message" : f"something went wrong \n {e}"})
    else:
         print("not post")
         print(request.method)
         return toJsonResponse({"status" : False,"message" : f"not post, something went wrong"})



@csrf_exempt
def create_catchment(request):
    try:
        if request.method == "POST":
            data = loadJsonData(request)
            name = data.get("name")
            exists = Catchment.objects.filter(name = name).exists()
            if exists:
                return toJsonResponse({"status" : False,"message" : f"Failed to create catechment, Maybe it already exists"})

            if name:
                catchment = Catchment(name = name)
                catchment.save()
                return toJsonResponse({"status" : True,"message" : f"catchment {name} created"})
            else:
                return toJsonResponse({"status" : False,"message" : f"Failed to create catechment, Maybe it already exists"})
    except Exception as e:
        print(e)
        return toJsonResponse({"status" : False,"message" : f"something went wrong"})
    



@csrf_exempt
def create_project(request):
    try:
        if request.method == "POST":
            data = loadJsonData(request)
            catchment = data.get("catchment")
            name = data.get("name")
            catchment_name = Catchment.objects.get(name = catchment)
            exists = Project.objects.filter(name = name).exists()
            if exists:
                return toJsonResponse({"status" : False,"message" : "Failed to create project, Maybe it already exists"})
            
            if catchment_name:
                    project_code = data.get("project_code")
                    Phase_Name = data.get("Phase_Name")
                    target_HHs = data.get("target_HHs")
                    enrolled_HHs = data.get("enrolled_HHs")
                    project_code = data.get("project_code")
                    new_project = Project(name = name,catchment = catchment_name,project_code=project_code,Phase_Name=Phase_Name,target_HHs=target_HHs,enrolled_HHs=enrolled_HHs)
                    new_project.save()
                    return toJsonResponse({"status" : True,"message" : f"Project {name} created"})

            else:
                return toJsonResponse({"status" : False,"message" : "Failed to create Project, Maybe it already exists"})
    except Exception as e:
        print(e)
        return toJsonResponse({"status" : False,"message" : "something went wrong"})

@csrf_exempt
def create_person(request):
    try:
        if request.method == "POST":
            data = loadJsonData(request)
            first_name = data.get("first_name")
            last_name = data.get("last_name")
            full_name = data.get("full_name")
            form_number = data.get("form_number")
            phoneNumber = data.get("phoneNumber")
            national_id = data.get("national_id")
            gender = data.get("gender")
            district_name = data.get("district_name")
            traditional_authority_name = data.get("traditional_authority_name")
            household_id = data.get("household_id")
            group_village_head_name = data.get("group_village_head_name")
            village_name = data.get("village_name")
            project = data.get("project")
            
            exists = Person.objects.filter(form_number = form_number).exists()
            if exists:
                return toJsonResponse({"status" : False,"message" : "Person with given form number already exists"})
            
            
            if form_number and first_name and last_name and project and Catchment:
                user_project = Project.objects.get(name = project)                
                person = Person(first_name = first_name,last_name = last_name,form_number = form_number,village_name=village_name,group_village_head_name=group_village_head_name,household_id=household_id,traditional_authority_name=traditional_authority_name,district_name = district_name,gender = gender,national_id=national_id,phoneNumber=phoneNumber,project = user_project,full_name=full_name)
                person.save()
                return toJsonResponse({"status" : True,"message" : f"Person {full_name} created"})  
    except Exception as e:
        print(e)
        return toJsonResponse({"status" : False,"message" : "something went wrong"})
    
def checkIfCSV(path : str) -> bool:
    stringList = path.split(".")
    return stringList[1] and stringList[1] == "csv"

def csv_to_xlsx(csvpath):
    my_workbook = Workbook()
    sheet = my_workbook.active
    text_content = StringIO(csvpath.getvalue().decode("utf-8"))
    reader = csv.reader(text_content)
    for row in reader:
            sheet.append(row)
    return my_workbook

@csrf_exempt
def add_excel_data(request):
    if request.method == "OPTIONS":
        response = HttpResponse()
        response['Access-Control-Allow-Origin'] = '*' 
        response['Access-Control-Allow-Methods'] = 'POST'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response
    elif request.method == "POST":
        data = json.loads(request.body.decode())
        path = data.get("path")
        name = data.get("name")
        xpath = download(path=path,name=name)
        ref = create_path(xpath)
        populate_referance_from_file(file_name=ref)
        return toJsonResponse({"status" : True,"message" : "Table updated"})
    else :
        return toJsonResponse({"status" : False,"message" : "invalid request"})

def populate_referance_from_file(file_name):
    try:
        data = load_external_reference(path=file_name)
        catchments = Catchment.objects.all()
        projects = Project.objects.all()
        people = Person.objects.all()
        form_numbers = [person.form_number for person in people]
        project_names = [project.name for project in projects]
        catchment_names = [catchment.name for catchment in catchments]

        ref_row = [cell.value for cell in data[0]]
        print(ref_row)
        header = {
        key: ref_row.index(key) if key in ref_row else -1 
        for key in [
            "district_name", 
            "village_name", 
            "traditional_authority_name", 
            "catchment_name", 
            "group_village_head_name", 
            "Phase_Name", 
            "project_code", 
            "project_name", 
            "target_HHs", 
            "enrolled_HHs", 
            "form_number", 
            "full_name", 
            "project_id", 
            "household_id", 
            "Phone Number"
        ]
    }

        for row in data:
            names= row[header["full_name"]].value.split(" ")
            first_name = names[0]
            last_name = names[1] if len(names) > 1 else ""
            form_number = str(row[header["form_number"]].value)
            catchment = str(row[header["catchment_name"]].value)
            project = str(row[header["project_name"]].value)
            if form_number not in form_numbers:
                if project in project_names:
                    my_project = Project.objects.get(name = project)
                    participant = Person(
                        first_name = first_name,
                        last_name = last_name,
                        full_name = row[header["full_name"]].value,
                        form_number = form_number,
                        phoneNumber = row[header["Phone Number"]].value,
                        traditional_authority_name = row[header["traditional_authority_name"]].value,
                        group_village_head_name = row[header["group_village_head_name"]].value,
                        village_name = row[header["village_name"]].value,
                        household_id = row[header["household_id"]].value,
                        project = my_project
                    )
                    participant.save()
                else:
                    my_catchment = Catchment.objects.get(name = catchment) if catchment in catchment_names else Catchment(name = row[header["catchment_name"]].value)
                    if catchment not in catchment_names:
                        my_catchment.save()
                        catchment_names.append(catchment)
                        my_catchment = Catchment.objects.get(name = catchment)
                    
                    my_project = Project.objects.get(name = project) if project in project_names else Project(name = row[header["project_name"]].value,catchment = my_catchment,project_code = row[header["project_code"]].value,Phase_Name = row[header["Phase_Name"]].value,target_HHs = row[header["target_HHs"]].value,enrolled_HHs = row[header["enrolled_HHs"]].value,project_id = row[header["project_id"]].value)
                    if project not in project_names:
                        my_project.save()
                        project_names.append(project)
                        my_project = Project.objects.get(name = project)
                    participant = Person(
                        first_name = first_name,
                        last_name = last_name,
                        full_name = row[header["full_name"]].value,
                        form_number = form_number,
                        phoneNumber = row[header["Phone Number"]].value,
                        traditional_authority_name = row[header["traditional_authority_name"]].value,
                        group_village_head_name = row[header["group_village_head_name"]].value,
                        village_name = row[header["village_name"]].value,
                        household_id = row[header["household_id"]].value,
                        project = my_project
                    )
                    participant.save()
                    form_numbers.append(form_number)
            else:
                user = Person.objects.get(form_number = form_number)
                user.phoneNumber = row[header["Phone Number"]].value
                user.save()
        return toJsonResponse({"status" : True, "message" : "operation complete"})
            

    except Exception as e:
        print(e)
        return toJsonResponse({"status" : True, "message" : f"something went wrong \n {e}"})


# def add_phone_numbers(target_path : str,reference : list,output_path : str = "output2.xlsx" )->workbook.workbook.Workbook:
#     my_workbook = load_workbook(filename = target_path)
#     my_rows = [rows for rows in my_workbook.active.iter_rows()]
#     header_target = [ref.value for ref in reference[0]].index("Phone Number")
#     header_formnumber = [ref.value for ref in reference[0]].index("form_number")
#     header_target_final = [ref.value for ref in my_rows[0]].index("Phone Number")
#     formnumber_target_final = [ref.value for ref in my_rows[0]].index("form_number")
    
#     if header_target and header_target_final:
#         for phone in enumerate(my_rows):
#             if phone[0] != 0:
#                 target_row = phone[1]
#                 target_cell = target_row[header_target_final]
#                 for value in reference:
#                     if str(value[header_formnumber].value) == str(target_row[formnumber_target_final].value):
#                         target_cell.value = value[header_target_final].value  
#                         print(1)
#                     else :
#                         print(f"{str(value[header_formnumber].value)}  == {str(target_row[formnumber_target_final].value)}")
            
        
#     my_workbook.save(output_path)
#     return my_workbook
#															



#   = models.CharField(max_length=30)
#      = models.CharField(max_length=30)
#      = models.CharField(max_length=30,primary_key=True)
#      = models.CharField(max_length=30, default="NONE")
#      = models.CharField(max_length=30, default="NONE")
#      = models.CharField(max_length=30, default="NONE")
#      = models.CharField(max_length=30, default="NONE")
#      = models.CharField(max_length=30, default="NONE")
#      = models.CharField(max_length=30, default="NONE")
#      = models.CharField(max_length=30, default="NONE")
#      = models.CharField(max_length=30, default="NONE")
#      = models.CharField(max_length=30, default="NONE")
#      = models.CharField(max_length=30, default="NONE")
#      = models.ForeignKey(Project,on_delete=models.CASCADE,related_name="person")   


